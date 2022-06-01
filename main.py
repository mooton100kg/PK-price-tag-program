import pandas as pd
from datetime import date
from openpyxl.styles import Border, Font, Alignment, Side
from openpyxl import load_workbook
from math import ceil

def print_code(print_paper,path):
    wb = load_workbook(path)
    ws_name = str(date.today())
    if ws_name in wb.sheetnames:
        ws_name = ws_name + '-no.' +str(len(wb.sheetnames))
    wb.create_sheet(ws_name)
    ws = wb[ws_name]
    for i, item in enumerate(print_paper):
        long = len(item)
        if long >= 17:
            F = 10
        elif long <17:
            F = 11
        col = ceil((i+1)/2)
        sett = ceil(col/5)
        if col%5 == 0:
            column = 5
        else:
            column = col%5
        if i % 2 == 0:         
            row = sett*2-1
            cell = ws.cell(row = row, column = column)
            ws.cell(column=column, row = row, value = item)
            cell.font = Font(size=F)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(top=Side(border_style='thin',
                                           color='00000000'),
                                 left=Side(border_style='thin',
                                           color='00000000'),
                                 right=Side(border_style='thin',
                                           color='00000000'))
        elif i % 2 == 1:
            row = sett*2
            cell = ws.cell(row = row, column = column)
            ws.cell(column=column, row = row, value = item)
            cell.font = Font(size=F)
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(bottom=Side(border_style='thin',
                                            color='00000000'),
                                 left=Side(border_style='thin',
                                            color='00000000'),
                                 right=Side(border_style='thin',
                                            color='00000000'))
    for i in range(1,6):
        ws.column_dimensions[chr(64+i)].width = 18.5
    ws.page_margins.top = 0.2
    ws.page_margins.bottom = 0.2
    ws.page_margins.right = 0.2
    ws.page_margins.left = 0.2
    ws.page_margins.footer = 0
    ws.page_margins.header = 0
    ws.page_setup.paperSize = 9
    wb.save(filename = path)
    
def code_convert(cost,sell,supplier,year,month,gen):
    sell_code = ('S','R','Y','Z','M','L','H','G','P','J')
    Cost_code = ('N','E','D','I','X','A','F','O','C','B')
    sc = ''
    s = [int(i) for i in str(sell)]
    for i in s:
        sc += sell_code[i-1]
    cc = ''
    c = [int(i) for i in str(cost)]
    for i in c:
        cc += Cost_code[i-1]
    yc = ''
    y = [int(i) for i in str(year)]
    for i in y:
        yc += Cost_code[i-1]
    mc = ''
    m = [int(i) for i in str(month)]
    for i in m:
        mc += Cost_code[i-1]
    
    code = 'K' + sc + 'W' + supplier + cc + yc + mc + gen
    return code

def round_up(n, decimals=0):
    multiplier = 10 ** decimals
    return ceil(n * multiplier) / multiplier

def sell_price_cal(Cost):
    Sell_Price = ceil(Cost*1.5*(10**-1))*10
    return Sell_Price
    
    

